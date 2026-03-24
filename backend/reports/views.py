from datetime import datetime
from pathlib import Path

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import MedicationAccessPermission
from medications.municipality_catalog import (
    ORDERED_MUNICIPALITY_NAMES,
    get_display_municipality_name,
)
from medications.models import Medication, Municipality, Movement, MunicipalityStock
from django.db.models import Case, IntegerField, Sum, When

MONTHS_ES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


def parse_month(value: str | None):
    if not value:
        now = timezone.now()
        return now.year, now.month
    try:
        parsed = datetime.strptime(value, "%Y-%m")
        return parsed.year, parsed.month
    except ValueError:
        return None, None


def format_period(year_value: int, month_value: int) -> str:
    month_name = MONTHS_ES[month_value - 1] if 1 <= month_value <= 12 else str(month_value)
    return f"{month_name} {year_value}"


def parse_medication_ids(value: str | None):
    if not value:
        return []
    ids: list[int] = []
    for chunk in str(value).split(","):
        part = chunk.strip()
        if not part:
            continue
        if not part.isdigit():
            raise ValueError("medication_ids debe ser una lista de enteros separada por coma.")
        number = int(part)
        if number <= 0:
            raise ValueError("medication_ids debe contener solo enteros positivos.")
        ids.append(number)
    return sorted(set(ids))


def get_report_municipality_names() -> list[str]:
    return list(ORDERED_MUNICIPALITY_NAMES)


def build_municipality_medication_report(municipality, year_value: int, month_value: int):
    medications = list(
        Medication.objects.order_by("material_name").values("id", "code", "material_name")
    )
    stock_map = {
        row["medication_id"]: row["total"] or 0
        for row in MunicipalityStock.objects.filter(municipality=municipality)
        .values("medication_id")
        .annotate(total=Sum("stock"))
    }
    movement_rows = (
        Movement.objects.filter(
            municipality=municipality,
            created_at__year=year_value,
            created_at__month=month_value,
        )
        .values("medication_id")
        .annotate(
            ingresos=Sum(
                Case(
                    When(type="ingreso", then="quantity"),
                    default=0,
                    output_field=IntegerField(),
                )
            ),
            egresos=Sum(
                Case(
                    When(type="egreso", then="quantity"),
                    default=0,
                    output_field=IntegerField(),
                )
            ),
        )
    )
    movement_map = {
        row["medication_id"]: {
            "ingresos": row["ingresos"] or 0,
            "egresos": row["egresos"] or 0,
        }
        for row in movement_rows
    }

    items = []
    total_ingresos = 0
    total_egresos = 0
    total_quantity = 0

    for medication in medications:
        medication_id = medication["id"]
        ingresos = movement_map.get(medication_id, {}).get("ingresos", 0)
        egresos = movement_map.get(medication_id, {}).get("egresos", 0)
        stock = stock_map.get(medication_id, 0)
        items.append(
            {
                "code": medication["code"],
                "material_name": medication["material_name"],
                "ingresos": ingresos,
                "egresos": egresos,
                "real_time_stock": stock,
            }
        )
        total_ingresos += ingresos
        total_egresos += egresos
        total_quantity += ingresos + egresos

    return {
        "items": items,
        "total_quantity": total_quantity,
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
    }


class MunicipalityMonthlyReportView(APIView):
    permission_classes = [IsAuthenticated, MedicationAccessPermission]

    def get(self, request):
        municipality_id = request.query_params.get("municipality_id")
        year = request.query_params.get("year")
        month = request.query_params.get("month")
        year_value, month_value = parse_month(month)

        if not municipality_id:
            return Response({"detail": "municipality_id es requerido."}, status=400)

        if year and month:
            return Response({"detail": "Usa solo month (YYYY-MM) o year/month."}, status=400)

        if not year_value or not month_value:
            if year and month:
                year_value = int(year)
                month_value = int(month)
            else:
                year_value, month_value = parse_month(None)

        try:
            municipality = Municipality.objects.get(pk=int(municipality_id))
        except (ValueError, Municipality.DoesNotExist):
            return Response({"detail": "Municipio invalido."}, status=400)

        report_data = build_municipality_medication_report(municipality, year_value, month_value)

        return Response(
            {
                "municipality_id": municipality.id,
                "municipality_name": municipality.name,
                "year": year_value,
                "month": month_value,
                "total_quantity": report_data["total_quantity"],
                "total_ingresos": report_data["total_ingresos"],
                "total_egresos": report_data["total_egresos"],
                "items": report_data["items"],
            }
        )


class MunicipalityMonthlyReportDownloadView(APIView):
    permission_classes = [IsAuthenticated, MedicationAccessPermission]

    def get(self, request):
        municipality_id = request.query_params.get("municipality_id")
        month = request.query_params.get("month")
        export_format = (request.query_params.get("export_format") or "pdf").lower()
        year_value, month_value = parse_month(month)

        if not municipality_id:
            return Response({"detail": "municipality_id es requerido."}, status=400)

        if str(municipality_id).lower() == "all":
            # Compatibilidad: permite descargar consolidado usando el endpoint
            # base /download/ con municipality_id=all.
            return AllMunicipalitiesMonthlyReportDownloadView().get(request)

        if not year_value or not month_value:
            year_value, month_value = parse_month(None)

        try:
            municipality = Municipality.objects.get(pk=int(municipality_id))
        except (ValueError, Municipality.DoesNotExist):
            return Response({"detail": "Municipio invalido."}, status=400)

        report_data = build_municipality_medication_report(municipality, year_value, month_value)
        if export_format == "excel":
            return self._build_excel(report_data, municipality, year_value, month_value, request)

        try:
            from io import BytesIO
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except Exception:
            return Response(
                {"detail": "Instala reportlab para generar PDF (pip install reportlab)."},
                status=500,
            )

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = []

        total_movements = report_data["total_quantity"]
        total_ingresos = report_data["total_ingresos"]
        total_egresos = report_data["total_egresos"]

        summary_data = [
            ["", f"Total de movimientos: {total_movements}", "", f"Ingresos: {total_ingresos}", "", f"Egresos: {total_egresos}"],
        ]
        summary_table = Table(summary_data, hAlign="CENTER", colWidths=[14, 170, 14, 120, 10, 100])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef3fb")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (2, -1), "LEFT"),
                    ("ALIGN", (3, 0), (3, -1), "LEFT"),
                    ("ALIGN", (4, 0), (4, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.HexColor("#1f4f9c")),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#1f4f9c")),
                    ("GRID", (0, 0), (-1, -1), 0.0, colors.white),
                ]
            )
        )
        elements.append(Spacer(1, 170))
        elements.append(summary_table)
        elements.append(Spacer(1, 10))

        data = [["No.", "Codigo", "Material medico", "Ingresos", "Egresos", "Existencia"]]
        row_index = 1
        for item in report_data["items"]:
            data.append(
                [
                    str(row_index),
                    item["code"],
                    item["material_name"],
                    str(item["ingresos"]),
                    str(item["egresos"]),
                    str(item["real_time_stock"]),
                ]
            )
            row_index += 1
        table = Table(data, hAlign="CENTER", colWidths=[28, 50, 220, 60, 60, 95])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4f9c")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfd8e6")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f6fb")]),
                    ("ALIGN", (0, 0), (1, -1), "CENTER"),
                    ("ALIGN", (2, 0), (2, -1), "LEFT"),
                    ("ALIGN", (3, 0), (5, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.HexColor("#1f4f9c")),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#1f4f9c")),
                ]
            )
        )
        elements.append(table)

        footer_style = styles["Normal"].clone("footer_style")
        footer_style.alignment = 1
        footer_style.fontSize = 9
        elements.append(Spacer(1, 18))

        elements.append(Paragraph("Reporte generado automaticamente por el sistema SISAS", footer_style))

        def draw_header(canvas_obj, doc_obj):
            canvas_obj.saveState()
            width, height = letter
            header_top = height - 40
            canvas_obj.setFillColor(colors.HexColor("#1f4f9c"))
            canvas_obj.rect(0, header_top - 90, width, 90, fill=1, stroke=0)
            canvas_obj.setFillColor(colors.white)
            canvas_obj.setFont("Helvetica-Bold", 16)
            canvas_obj.drawCentredString(width / 2, header_top - 40, "REPORTE QUINCENAL DE")
            canvas_obj.setFont("Helvetica-Bold", 20)
            canvas_obj.drawCentredString(width / 2, header_top - 65, "INSUMOS / REACTIVOS")

            canvas_obj.setFillColor(colors.HexColor("#0f2c5c"))
            canvas_obj.setFont("Helvetica-Bold", 12)
            canvas_obj.drawCentredString(
                width / 2,
                height - 15,
                "DIRECCION DEPARTAMENTAL DE REDES INTEGRADAS DE SERVICIOS DE SALUD",
            )

            # Info box
            canvas_obj.setFillColor(colors.HexColor("#eef3fb"))
            info_box_top = header_top - 105
            canvas_obj.roundRect(60, info_box_top - 45, width - 120, 45, 8, fill=1, stroke=0)
            canvas_obj.setFillColor(colors.HexColor("#0f2c5c"))
            canvas_obj.setFont("Helvetica-Bold", 10)
            date_label = timezone.localdate().strftime("%d/%m/%Y")
            username = request.user.get_full_name() or request.user.username
            canvas_obj.drawString(80, info_box_top - 20, f"DMS/RED LOCAL: {municipality.name}")
            canvas_obj.drawString(width / 2 + 10, info_box_top - 20, f"Fecha: {date_label}")
            canvas_obj.drawString(80, info_box_top - 35, f"Usuario: {username}")

            # Summary pills
            total_movements = report_data["total_quantity"]
            total_ingresos = report_data["total_ingresos"]
            total_egresos = report_data["total_egresos"]
            canvas_obj.setFillColor(colors.HexColor("#eef3fb"))
            # summary moved below table

            # Footer
            canvas_obj.setFont("Helvetica", 9)
            canvas_obj.setFillColor(colors.HexColor("#5f6b7a"))
            canvas_obj.drawCentredString(width / 2, 30, "Direccion de Area de Salud de Sololá | 2026")
            canvas_obj.restoreState()

        doc.build(elements, onFirstPage=draw_header)

        pdf = buffer.getvalue()
        buffer.close()

        filename = f"reporte_{municipality.name}_{year_value}-{month_value:02d}.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _build_excel(self, report_data, municipality, year_value, month_value, request):
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:
            return Response(
                {"detail": "Instala openpyxl para generar EXCEL (pip install openpyxl)."},
                status=500,
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        header_fill = PatternFill(start_color="1F4F9C", end_color="1F4F9C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(bold=True)
        centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

        start_col_idx = 3
        cols = [get_column_letter(start_col_idx + i) for i in range(6)]
        title_row_1 = 5
        title_row_2 = 6
        dms_row = 7
        date_row = 8
        user_row = 9
        table_header_row = 10

        for row in [title_row_1, title_row_2, dms_row, date_row, user_row]:
            ws.merge_cells(f"{cols[0]}{row}:{cols[-1]}{row}")

        ws[f"{cols[0]}{title_row_1}"] = "DIRECCION DEPARTAMENTAL DE REDES INTEGRADAS DE SERVICIOS DE SALUD"
        ws[f"{cols[0]}{title_row_2}"] = "REPORTE QUINCENAL DE INSUMOS / REACTIVOS"
        ws[f"{cols[0]}{dms_row}"] = f"DMS/RED LOCAL: {municipality.name}"
        ws[f"{cols[0]}{date_row}"] = f"Fecha: {timezone.localdate().strftime('%d/%m/%Y')}"
        username = request.user.get_full_name() or request.user.username
        ws[f"{cols[0]}{user_row}"] = f"Usuario: {username}"

        for row in [title_row_1, title_row_2, dms_row, date_row, user_row]:
            ws[f"{cols[0]}{row}"].alignment = centered
        ws[f"{cols[0]}{title_row_1}"].font = title_font
        ws[f"{cols[0]}{title_row_2}"].font = title_font

        headers = ["No.", "Codigo", "Material medico", "Ingresos", "Egresos", "Existencia"]
        for idx, value in enumerate(headers):
            cell = ws.cell(row=table_header_row, column=start_col_idx + idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered

        row_number = 1
        for item in report_data["items"]:
            values = [
                row_number,
                item["code"],
                item["material_name"],
                item["ingresos"],
                item["egresos"],
                item["real_time_stock"],
            ]
            row_idx = ws.max_row + 1
            for idx, value in enumerate(values):
                ws.cell(row=row_idx, column=start_col_idx + idx, value=value).alignment = centered
            row_number += 1

        widths = {"C": 7, "D": 12, "E": 34, "F": 12, "G": 12, "H": 22}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"reporte_{municipality.name}_{year_value}-{month_value:02d}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class AllMunicipalitiesMonthlyReportDownloadView(APIView):
    permission_classes = [IsAuthenticated, MedicationAccessPermission]

    def get(self, request):
        month = request.query_params.get("month")
        export_format = (request.query_params.get("export_format") or "pdf").lower()
        medication_ids_param = request.query_params.get("medication_ids")
        year_value, month_value = parse_month(month)
        if not year_value or not month_value:
            year_value, month_value = parse_month(None)

        try:
            medication_ids = parse_medication_ids(medication_ids_param)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        movements = (
            Movement.objects.filter(
                created_at__year=year_value,
                created_at__month=month_value,
            )
            .select_related("medication", "user", "municipality")
            .order_by("municipality__name", "created_at", "id")
        )
        if medication_ids:
            movements = movements.filter(medication_id__in=medication_ids)

        rows = []
        for index, movement in enumerate(movements, start=1):
            user_name = movement.user.get_full_name() or movement.user.username if movement.user else "-"
            rows.append(
                [
                    index,
                    movement.municipality.name if movement.municipality else "-",
                    movement.medication.code,
                    movement.medication.material_name,
                    "Ingreso" if movement.type == "ingreso" else "Egreso",
                    user_name,
                    movement.quantity,
                ]
            )

        if export_format == "excel":
            return self._build_excel(rows, year_value, month_value, request, medication_ids)
        return self._build_pdf(movements, year_value, month_value, request, medication_ids)

    def _build_excel(self, rows, year_value: int, month_value: int, request, medication_ids=None):
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:
            return Response(
                {"detail": "Instala openpyxl para generar EXCEL (pip install openpyxl)."},
                status=500,
            )

        from django.db.models import Sum

        wb = Workbook()

        header_fill = PatternFill(start_color="1F4F9C", end_color="1F4F9C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_font = Font(bold=True)

        def style_header(sheet, row_index: int, col_letters: list[str]):
            for col in col_letters:
                cell = sheet[f"{col}{row_index}"]
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        def apply_table_format(sheet, header_row: int, col_letters: list[str]):
            last_row = sheet.max_row
            min_col = ord(col_letters[0]) - ord("A") + 1
            max_col = ord(col_letters[-1]) - ord("A") + 1
            for row in sheet.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.alignment = centered_alignment

            for col in col_letters:
                sheet[f"{col}{header_row}"].alignment = header_alignment

            sheet.row_dimensions[header_row].height = 22
            for row_idx in range(header_row + 1, last_row + 1):
                sheet.row_dimensions[row_idx].height = 20

            sheet.auto_filter.ref = f"{col_letters[0]}{header_row}:{col_letters[-1]}{last_row}"
            sheet.freeze_panes = f"{col_letters[0]}{header_row + 1}"
            sheet.print_options.horizontalCentered = True

        selected_medications = Medication.objects.order_by("material_name")
        if medication_ids:
            selected_medications = selected_medications.filter(id__in=medication_ids)
        medication_items = list(selected_medications.values_list("id", "material_name"))

        municipality_display_by_id = {
            municipality.id: get_display_municipality_name(municipality.name)
            for municipality in Municipality.objects.all()
        }
        ordered_municipality_names = get_report_municipality_names()

        # Detail data maps
        downloaded_by = request.user.get_full_name() or request.user.username
        movements = Movement.objects.filter(
            created_at__year=year_value,
            created_at__month=month_value,
            municipality__isnull=False,
        ).select_related("municipality", "medication")
        if medication_ids:
            movements = movements.filter(medication_id__in=medication_ids)

        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        egresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="egreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        stock_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in MunicipalityStock.objects.filter(
                medication_id__in=medication_ids
            ).values("municipality__id", "municipality__name").annotate(total=Sum("stock"))
        } if medication_ids else {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in MunicipalityStock.objects.values("municipality__id", "municipality__name")
            .annotate(total=Sum("stock"))
        }

        # Sheet 1: General summary
        start_col_idx = 3  # C
        start_col = get_column_letter(start_col_idx)
        general_cols = [get_column_letter(start_col_idx + i) for i in range(6)]  # C:H
        title_row_1 = 5
        title_row_2 = 6
        dms_row = 7
        date_row = 8
        user_row = 9
        table_header_row = 10

        ws = wb.active
        ws.title = "Detalle general"
        ws.merge_cells(f"{general_cols[0]}{title_row_1}:{general_cols[-1]}{title_row_1}")
        ws.merge_cells(f"{general_cols[0]}{title_row_2}:{general_cols[-1]}{title_row_2}")
        ws.merge_cells(f"{general_cols[0]}{dms_row}:{general_cols[-1]}{dms_row}")
        ws.merge_cells(f"{general_cols[0]}{date_row}:{general_cols[-1]}{date_row}")
        ws.merge_cells(f"{general_cols[0]}{user_row}:{general_cols[-1]}{user_row}")

        ws[f"{general_cols[0]}{title_row_1}"] = "DIRECCION DEPARTAMENTAL DE REDES INTEGRADAS DE SERVICIOS DE SALUD"
        ws[f"{general_cols[0]}{title_row_2}"] = "REPORTE QUINCENAL DE INSUMOS / REACTIVOS"
        ws[f"{general_cols[0]}{dms_row}"] = "DMS/RED LOCAL: CONSOLIDADO GENERAL"
        ws[f"{general_cols[0]}{date_row}"] = f"Fecha: {timezone.localdate().strftime('%d/%m/%Y')}"
        ws[f"{general_cols[0]}{user_row}"] = f"Usuario: {downloaded_by}"

        ws[f"{general_cols[0]}{title_row_1}"].font = title_font
        ws[f"{general_cols[0]}{title_row_2}"].font = title_font
        ws[f"{general_cols[0]}{title_row_1}"].alignment = centered_alignment
        ws[f"{general_cols[0]}{title_row_2}"].alignment = centered_alignment
        ws[f"{general_cols[0]}{dms_row}"].alignment = centered_alignment
        ws[f"{general_cols[0]}{date_row}"].alignment = centered_alignment
        ws[f"{general_cols[0]}{user_row}"].alignment = centered_alignment

        for col_idx, value in enumerate(["No.", "DMS/RED LOCAL", "Insumo", "Ingresos", "Salidas (Egresos)", "Existencia"]):
            ws.cell(row=table_header_row, column=start_col_idx + col_idx, value=value)
        style_header(ws, table_header_row, general_cols)

        from django.db.models import Sum, Case, When, IntegerField

        movements_summary = movements.values("municipality_id", "medication_id").annotate(
            ingresos=Sum(Case(When(type="ingreso", then="quantity"), default=0, output_field=IntegerField())),
            egresos=Sum(Case(When(type="egreso", then="quantity"), default=0, output_field=IntegerField())),
        )
        movement_map: dict[tuple[str, int], tuple[int, int]] = {}
        for row in movements_summary:
            municipality_name = municipality_display_by_id.get(row["municipality_id"])
            if not municipality_name:
                continue
            movement_map[(municipality_name, row["medication_id"])] = (
                row["ingresos"] or 0,
                row["egresos"] or 0,
            )

        stock_rows = MunicipalityStock.objects.values("municipality_id", "medication_id").annotate(total=Sum("stock"))
        if medication_ids:
            stock_rows = stock_rows.filter(medication_id__in=medication_ids)

        stock_map: dict[tuple[str, int], int] = {}
        for row in stock_rows:
            municipality_name = municipality_display_by_id.get(row["municipality_id"])
            if not municipality_name:
                continue
            stock_map[(municipality_name, row["medication_id"])] = row["total"] or 0

        row_number = 1
        for municipality_name in ordered_municipality_names:
            for medication_id, medication_name in medication_items:
                ingresos_total, egresos_total = movement_map.get((municipality_name, medication_id), (0, 0))
                stock_total = stock_map.get((municipality_name, medication_id), 0)
                row_idx = ws.max_row + 1
                values = [row_number, municipality_name, medication_name, ingresos_total, egresos_total, stock_total]
                for col_idx, value in enumerate(values):
                    ws.cell(row=row_idx, column=start_col_idx + col_idx, value=value)
                row_number += 1

        ws.column_dimensions["C"].width = 7
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 34
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 12
        apply_table_format(ws, table_header_row, general_cols)

        # One sheet per municipality with summary
        for municipality_name in ordered_municipality_names:
            # Excel sheet title max length 31 and no invalid chars
            safe_title = "".join(ch for ch in municipality_name if ch not in '\\/*?:[]')
            safe_title = safe_title[:31] if safe_title else municipality_name
            sheet = wb.create_sheet(title=safe_title)
            muni_start_col_idx = 3  # C
            muni_cols = [get_column_letter(muni_start_col_idx + i) for i in range(4)]  # C:F
            muni_title_row_1 = 5
            muni_title_row_2 = 6
            muni_name_row = 7
            muni_date_row = 8
            muni_user_row = 9
            muni_table_header_row = 10

            sheet.merge_cells(f"{muni_cols[0]}{muni_title_row_1}:{muni_cols[-1]}{muni_title_row_1}")
            sheet.merge_cells(f"{muni_cols[0]}{muni_title_row_2}:{muni_cols[-1]}{muni_title_row_2}")
            sheet.merge_cells(f"{muni_cols[0]}{muni_name_row}:{muni_cols[-1]}{muni_name_row}")
            sheet.merge_cells(f"{muni_cols[0]}{muni_date_row}:{muni_cols[-1]}{muni_date_row}")
            sheet.merge_cells(f"{muni_cols[0]}{muni_user_row}:{muni_cols[-1]}{muni_user_row}")

            sheet[f"{muni_cols[0]}{muni_title_row_1}"] = "DIRECCION DEPARTAMENTAL DE REDES INTEGRADAS DE SERVICIOS DE SALUD"
            sheet[f"{muni_cols[0]}{muni_title_row_2}"] = "REPORTE QUINCENAL DE INSUMOS / REACTIVOS"
            sheet[f"{muni_cols[0]}{muni_name_row}"] = f"DMS/RED LOCAL: {municipality_name}"
            sheet[f"{muni_cols[0]}{muni_date_row}"] = f"Fecha: {timezone.localdate().strftime('%d/%m/%Y')}"
            sheet[f"{muni_cols[0]}{muni_user_row}"] = f"Usuario: {downloaded_by}"

            sheet[f"{muni_cols[0]}{muni_title_row_1}"].font = title_font
            sheet[f"{muni_cols[0]}{muni_title_row_2}"].font = title_font
            sheet[f"{muni_cols[0]}{muni_title_row_1}"].alignment = centered_alignment
            sheet[f"{muni_cols[0]}{muni_title_row_2}"].alignment = centered_alignment
            sheet[f"{muni_cols[0]}{muni_name_row}"].alignment = centered_alignment
            sheet[f"{muni_cols[0]}{muni_date_row}"].alignment = centered_alignment
            sheet[f"{muni_cols[0]}{muni_user_row}"].alignment = centered_alignment

            for col_idx, value in enumerate(["Insumo", "Ingresos", "Salidas (Egresos)", "Existencia"]):
                sheet.cell(row=muni_table_header_row, column=muni_start_col_idx + col_idx, value=value)
            style_header(sheet, muni_table_header_row, muni_cols)
            for medication_id, medication_name in medication_items:
                ingresos_total, egresos_total = movement_map.get((municipality_name, medication_id), (0, 0))
                stock_total = stock_map.get((municipality_name, medication_id), 0)
                row_idx = sheet.max_row + 1
                values = [medication_name, ingresos_total, egresos_total, stock_total]
                for col_idx, value in enumerate(values):
                    sheet.cell(row=row_idx, column=muni_start_col_idx + col_idx, value=value)
            sheet.column_dimensions["C"].width = 34
            sheet.column_dimensions["D"].width = 12
            sheet.column_dimensions["E"].width = 16
            sheet.column_dimensions["F"].width = 12
            apply_table_format(sheet, muni_table_header_row, muni_cols)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"reporte_todos_municipios_{year_value}-{month_value:02d}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _build_pdf(self, movements, year_value: int, month_value: int, request, medication_ids=None):
        try:
            from io import BytesIO
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except Exception:
            return Response(
                {"detail": "Instala reportlab para generar PDF (pip install reportlab)."},
                status=500,
            )

        from django.db.models import Sum

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=20, bottomMargin=20, leftMargin=36, rightMargin=36)
        styles = getSampleStyleSheet()
        elements = []

        selected_medications = Medication.objects.order_by("material_name")
        if medication_ids:
            selected_medications = selected_medications.filter(id__in=medication_ids)
        medication_items = list(selected_medications.values_list("id", "material_name"))

        municipality_display_by_id = {
            municipality.id: get_display_municipality_name(municipality.name)
            for municipality in Municipality.objects.all()
        }
        ordered_municipality_names = get_report_municipality_names()

        username = request.user.get_full_name() or request.user.username
        date_label = timezone.localdate().strftime("%d/%m/%Y")

        total_movements = movements.count()
        total_ingresos = movements.filter(type="ingreso").count()
        total_egresos = movements.filter(type="egreso").count()

        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        ingresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="ingreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        egresos_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in movements.filter(type="egreso")
            .values("municipality__id", "municipality__name")
            .annotate(total=Sum("quantity"))
        }
        stock_map = {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in MunicipalityStock.objects.filter(
                medication_id__in=medication_ids
            ).values("municipality__id", "municipality__name").annotate(total=Sum("stock"))
        } if medication_ids else {
            (row["municipality__id"], row["municipality__name"]): row["total"] or 0
            for row in MunicipalityStock.objects.values("municipality__id", "municipality__name")
            .annotate(total=Sum("stock"))
        }

        # Map of municipality -> list of insumos (material_name) for the month
        materials_map: dict[tuple[int | None, str | None], list[str]] = {}
        for row in movements.values(
            "municipality__id", "municipality__name", "medication__material_name"
        ).distinct():
            key = (row["municipality__id"], row["municipality__name"])
            materials_map.setdefault(key, []).append(row["medication__material_name"])

        def format_materials(key: tuple[int | None, str | None], limit: int = 5) -> str:
            items = sorted(materials_map.get(key, []))
            if not items:
                return "-"
            if len(items) <= limit:
                return ", ".join(items)
            extra = len(items) - limit
            return f"{', '.join(items[:limit])} (+{extra})"

        from django.db.models import Sum, Case, When, IntegerField

        movements = movements.filter(municipality__isnull=False)
        movements_summary = movements.values("municipality_id", "medication_id").annotate(
            ingresos=Sum(Case(When(type="ingreso", then="quantity"), default=0, output_field=IntegerField())),
            egresos=Sum(Case(When(type="egreso", then="quantity"), default=0, output_field=IntegerField())),
        )
        movement_map: dict[tuple[str, int], tuple[int, int]] = {}
        for row in movements_summary:
            municipality_name = municipality_display_by_id.get(row["municipality_id"])
            if not municipality_name:
                continue
            movement_map[(municipality_name, row["medication_id"])] = (
                row["ingresos"] or 0,
                row["egresos"] or 0,
            )

        stock_rows = MunicipalityStock.objects.values("municipality_id", "medication_id").annotate(total=Sum("stock"))
        if medication_ids:
            stock_rows = stock_rows.filter(medication_id__in=medication_ids)

        stock_map: dict[tuple[str, int], int] = {}
        for row in stock_rows:
            municipality_name = municipality_display_by_id.get(row["municipality_id"])
            if not municipality_name:
                continue
            stock_map[(municipality_name, row["medication_id"])] = row["total"] or 0

        data = [["No.", "DMS/RED LOCAL", "Insumo", "Ingresos", "Salidas (Egresos)", "Existencia"]]
        row_number = 1
        for municipality_name in ordered_municipality_names:
            for medication_id, medication_name in medication_items:
                ingresos_total, egresos_total = movement_map.get((municipality_name, medication_id), (0, 0))
                stock_total = stock_map.get((municipality_name, medication_id), 0)
                data.append(
                    [
                        str(row_number),
                        municipality_name,
                        medication_name,
                        str(ingresos_total),
                        str(egresos_total),
                        str(stock_total),
                    ]
                )
                row_number += 1

        table = Table(data, hAlign="CENTER", colWidths=[30, 160, 240, 70, 90, 90])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4f9c")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfd8e6")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f6fb")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (2, 0), (2, -1), "LEFT"),
                    ("ALIGN", (3, 0), (5, -1), "RIGHT"),
                ]
            )
        )
        elements.append(table)

        def draw_header(canvas_obj, doc_obj):
            canvas_obj.saveState()
            width, height = landscape(letter)

            # Top light bar
            canvas_obj.setFillColor(colors.HexColor("#dcecff"))
            canvas_obj.rect(0, height - 30, width, 30, fill=1, stroke=0)
            canvas_obj.setFont("Helvetica-Bold", 10)
            canvas_obj.setFillColor(colors.HexColor("#0f2c5c"))
            canvas_obj.drawCentredString(width / 2, height - 20, "DIRECCION DEPARTAMENTAL DE REDES INTEGRADAS DE SERVICIOS DE SALUD")

            # Main title bar
            canvas_obj.setFillColor(colors.HexColor("#1f4f9c"))
            canvas_obj.rect(0, height - 95, width, 55, fill=1, stroke=0)
            canvas_obj.setFillColor(colors.white)
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawCentredString(width / 2, height - 60, "REPORTE QUINCENAL DE")
            canvas_obj.drawCentredString(width / 2, height - 78, "INSUMOS / REACTIVOS")

            # Info box
            canvas_obj.setFillColor(colors.HexColor("#eef3fb"))
            info_top = height - 115
            info_box_width = width - 120
            info_box_x = (width - info_box_width) / 2
            canvas_obj.roundRect(info_box_x, info_top - 45, info_box_width, 45, 8, fill=1, stroke=0)
            canvas_obj.setFillColor(colors.HexColor("#0f2c5c"))
            canvas_obj.setFont("Helvetica-Bold", 9)
            left_x = info_box_x + 20
            right_x = info_box_x + info_box_width / 2 + 20
            canvas_obj.drawString(left_x, info_top - 20, "DMS/RED LOCAL: CONSOLIDADO GENERAL")
            canvas_obj.drawString(right_x, info_top - 20, f"Fecha: {date_label}")
            canvas_obj.drawString(left_x, info_top - 35, f"Usuario: {username}")

            # Summary pills
            pill_top = info_top - 58
            pill_height = 18
            pill_widths = [180, 120, 120]
            pill_labels = [
                f"Total de movimientos: {total_movements}",
                f"Ingresos: {total_ingresos}",
                f"Egresos: {total_egresos}",
            ]
            total_pills_width = sum(pill_widths) + (len(pill_widths) - 1) * 14
            x = (width - total_pills_width) / 2
            for label, w in zip(pill_labels, pill_widths):
                canvas_obj.setFillColor(colors.HexColor("#eef3fb"))
                canvas_obj.roundRect(x, pill_top - pill_height, w, pill_height, 6, fill=1, stroke=0)
                canvas_obj.setFillColor(colors.HexColor("#0f2c5c"))
                canvas_obj.setFont("Helvetica-Bold", 8)
                canvas_obj.drawCentredString(x + w / 2, pill_top - 12, label)
                x += w + 14

            canvas_obj.restoreState()

        # Leave space for header block
        elements.insert(0, Spacer(1, 175))

        doc.build(elements, onFirstPage=draw_header)
        pdf = buffer.getvalue()
        buffer.close()

        filename = f"reporte_todos_municipios_{year_value}-{month_value:02d}.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
